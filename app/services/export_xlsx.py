from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from app.config import DOCS_DIR
def export_spec_xlsx(estimate, number: str) -> str:
    wb = Workbook(); ws = wb.active; ws.title = "Спецификация"
    ws.append(["Раздел","Поз.","Наименование","Ед.","Кол-во","Цена (мат.)","Цена (монтаж)","Сумма"])
    sec_map = {s.id: s for s in estimate.sections}
    for it in estimate.items:
        sec = sec_map.get(it.section_id); total = (it.qty or 0)*((it.price_material or 0)+(it.price_labor or 0))
        ws.append([sec.title if sec else "", it.position or "", it.name, it.unit or "", it.qty or 0, it.price_material or 0, it.price_labor or 0, total])
    for c in range(1,9): ws.column_dimensions[get_column_letter(c)].width = 18
    path = DOCS_DIR / f"Спецификация_{number}.xlsx"; wb.save(path); return str(path)
